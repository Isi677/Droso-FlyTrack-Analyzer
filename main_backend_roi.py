import pandas as pd
import os
from openpyxl import load_workbook

def obtain_mmpx_mmpy(excel_file):
    #El archivo excel se transforma csv                
    df1 = pd.read_excel(excel_file, sheet_name = 0) 
    name_file = excel_file.split(".")
    name_file = name_file[0]
    df1.to_csv(f"{name_file}.csv", index = None)

    #Se abre el csv creado para leer cada fila
    with open(f"{name_file}.csv", "r") as file:
        filas = file.readlines()
        #Se obtenien los datos de mm por pixel
        mm_px = filas[4].strip().split(",")
        mm_py = filas[5].strip().split(",")
        mm_px = float(mm_px[1])
        mm_py = float(mm_py[1])
    os.remove(f"{name_file}.csv")

    return mm_px, mm_py

def obtain_mmpx_mmpy_Bonsai(excel_file):
    name_file = excel_file.split(".")
    name_file = name_file[0]
    try:
        df = pd.read_excel(excel_file, sheet_name="Data ROI + mm per pixel")
    except:
        return None
    df.to_csv(f"{name_file}.csv", index = None)
    with open(f"{name_file}.csv", "r") as file:
        filas = file.readlines()
        #Se obtenien los datos de mm por pixel
        data_mmp = filas[1].strip().split(",")
        mm_px = float(data_mmp[4])
        mm_py = float(data_mmp[5])
    os.remove(f"{name_file}.csv")
    return mm_px, mm_py
    
def add_arenaroi_to_csvdata(length, roi_coordinates_dict):
    for i in range (0, length):
        csv_file = roi_coordinates_dict[i][1]
        coords = roi_coordinates_dict[i][2]
        diameter = roi_coordinates_dict[i][3]*2

        data_roi = coords
        mmpx = diameter/(abs(coords[0]-coords[2]))
        mmpy = diameter/(abs(coords[1]-coords[3]))
        data_roi.append(mmpx)
        data_roi.append(mmpy)
        data_roi.append(roi_coordinates_dict[i][0])
        df_roi = pd.DataFrame([data_roi], columns=["x0", "y0", "x1", "y1", "Xs (mm/px)", "Ys (mm/px)", "MP4 File"])

        with open(csv_file, "r") as file:
            filas = file.readlines()
         
        data = []  
        for j in range(0, len(filas)):
            fila = filas[j].strip().split(",")
            data.append(fila)
        df_data = pd.DataFrame(data)

        new_path = csv_file[:-4]+".xlsx"

        # Load the existing workbook if it exists
        try:
            workbook = load_workbook(new_path)
            # Remove the sheet if it already exists
            if "Data ROI + mm per pixel" in workbook.sheetnames:
                std = workbook["Data ROI + mm per pixel"]
                workbook.remove(std)
            if "Data Coordinates" in workbook.sheetnames:
                std = workbook["Data Coordinates"]
                workbook.remove(std)
            writer = pd.ExcelWriter(new_path, engine='openpyxl')
            writer.book = workbook
        except FileNotFoundError:
            # Create a new workbook if the file does not exist
            writer = pd.ExcelWriter(new_path, engine='openpyxl')
            
        # Write the new DataFrame to the specific sheet
        df_data.to_excel(writer, sheet_name="Data Coordinates", index=False, header=False)
        df_roi.to_excel(writer, sheet_name="Data ROI + mm per pixel", index=False)
        writer.save()

def add_roi_to_csvdata(length, roi_coordinates_dict):
    for i in range (0, length):
        excel_file = roi_coordinates_dict[i][1]
        coords = roi_coordinates_dict[i][2] #List of lists containing the coordinates of each object
        identities = roi_coordinates_dict[i][3]
        data = []
        for i in range (0, len(coords)):
            data.append(coords[i] + [identities[i]])
        df_roi = pd.DataFrame(data, columns=["x0", "y0", "x1", "y1", "Type of Object"])
        print(data)

        # Load the existing workbook if it exists
        try:
            workbook = load_workbook(excel_file)
            # Remove the sheet if it already exists
            if "Data Objects Coordinates" in workbook.sheetnames:
                std = workbook["Data Objects Coordinates"]
                workbook.remove(std)
            writer = pd.ExcelWriter(excel_file, engine='openpyxl')
            writer.book = workbook
        except FileNotFoundError:
            # Create a new workbook if the file does not exist
            writer = pd.ExcelWriter(excel_file, engine='openpyxl')

        # Write the new DataFrame to the specific sheet
        df_roi.to_excel(writer, sheet_name="Data Objects Coordinates", index=False)
        writer.save()